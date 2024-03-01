"""
snowpit_datasheet_parser.py

functions that enables efficient extraction of snowpit information from the standard ECCC spreadsheets
J. King ECCC/CRD/CPS 2021
M. Brady ECCC/CRD/CPS 2021
B. Montpetit ECCC/CRD/CPS 2024

"""
from io import BytesIO
from pathlib import Path

import numpy as np
import pandas as pd
from dateutil.parser import parse
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from shapely.geometry import Point
from shapely.wkt import dumps


# constants for metadata look-up tables
PIT_LUT = {
    'Location': 'B2', 'Surveyors': 'M2', 'Site': 'B4',
    'Latitude': 'H2', 'Longitude': 'H4', 'Pit_ID': 'B6',
    'Total_Depth': 'F6', 'UTM_Zone': 'H6', 'Slope': 'K6',
    'Date': 'M6', 'Time': 'O6', 'Notes': 'R1'
}
META_LUT = {
    'Location': 'A2', 'Surveyors': 'E2', 'Site': 'A4',
    'Time': 'E4', 'Pit_ID': 'A6', 'Date': 'E6', 
    'Weather': 'I2', 
    'Precip': ['J6', 'K6', 'L6', 'N6', 'P6'],           # circle one
    'Sky': ['J7', 'K7', 'L7', 'N7', 'P7'],              # circle one
    'Wind_1': ['J8', 'K8', 'L8', 'N8', 'P8'],           # circle one
    'Wind_2': ['J9', 'K9', 'L9'],                       # circle one
    'GroundCondition': ['K10', 'L10', 'N10'],           # circle one
    'SoilMoisture': ['K11', 'L11', 'N11', 'P11'],       # circle one
    'GroundRoughness': ['K12', 'L12', 'N12'],           # circle one
    'GroundVegetation': ['K13', 'L13', 'N13', 'P13'],   # circle one or more
    'HeightOfGroundVegetation': ['L14', 'N14', 'P14'],  # could be all 3
    'TreeCanopy': ['K15', 'L15', 'N15', 'P15'],         # circle one
}
# need these for meta property... 
PRECIP = ['None', 'Light Snowfall', 'Moderate Snowfall', 'Heavy Snowfall', 'Rain']
SKY = ['Clear', 'PC 25%', 'PC 50%', 'MC 75%', 'Overcast']
WIND_1 = ['Calm', 'Breeze', 'Light', 'Moderate', 'Strong']
WIND_2 = ['Intermittent', 'Gusty', 'Steady']
GRNDCON = ['Frozen', 'Thawed', 'Uncertain']
SOILMOIST = ['Dry', 'Moist', 'Wet', 'Saturated']
GRNDROUGH = ['Smooth', 'Rough', 'Rugged']
GRNDVEG = ['Bare', 'Grass', 'Shrub', 'Deadfall']
TREECANOPY = ['No Trees', 'Sparse (5-20%)', 'Open (20-70%)', 'Closed (>70%)']


class SnowPitSheet():
    '''
    Python representation of a typical SnowPit Excel document

    usage: feed in a path (or Pathlike Object) to the `excel_file` argument, and as long
    as the excel doc is formatted similarly to Pit_DDMMYY_SS_PP_V2.xlsx then the 
    metadata/pit data will be extracted.

    object properties: 
        `meta`: dict of metadata describing the pit (extracted from sheet cells)
        `pit`: pandas dataframe multi-indexed by levels: 
            `density`
            `temperature`
            `stratigraphy`
            `ssa`
    '''
    def __init__(self, excel_file):
        if not isinstance(excel_file, Path):
            excel_file = Path(excel_file)
        if not excel_file.exists():
            raise FileNotFoundError('No such file: %s' % str(excel_file.absolute()))
            return
        elif not excel_file.suffix in ['.xls', '.xlsx']:
            raise ValueError('File is not an Excel workbook: %s' % str(excel_file.absolute()))
            return
        if excel_file.suffix == '.xls':
            raise TypeError('Old format (.xls) not supported. Please convert %s to %s.xlsx' % (str(excel_file), excel_file.stem))
            return
        # data-only forces any Excel cell-functions to return their result value
        try:
            # hack to get around openpyxl not properly closing a file
            with excel_file.open('rb') as f:
                in_mem_data = BytesIO(f.read())
            wb = load_workbook(in_mem_data, read_only=True, data_only=True)
        except:
            raise IOError('Error reading %s' % str(excel_file.absolute()))
            return
        self.source_file = excel_file.name
        self.meta = wb
        pit_sheet = wb['PIT']
        self.data = pit_sheet
        wb.close()

    def __str__(self):
        summary = f"{self.meta['location']} | " + \
            f"Timestamp: {self.meta['timestamp']} | " + \
            f"Data Rows: {len(self.data['ssa'])} | " + \
            f"Site: {self.meta['site']}"
        return summary

    @property
    def meta(self):
        return self.__meta

    @meta.setter
    def meta(self, workbook):
        if not isinstance(workbook, Workbook):
            raise TypeError('May only overwrite `meta` property with another SnowPit excel document opened as an openpyxl Workbook')
            return
        self.__meta = extract_metadata(workbook)

    @property
    def data(self):
        return self.__data

    @data.setter
    def data(self, sheet):
        if not isinstance(sheet, ReadOnlyWorksheet):
            raise TypeError('May only overwrite `pit` property with another SnowPit excel sheet opened as an openpyxl ReadOnlyWorksheet')
            return
        self.__data = extract_pit_data(sheet)


def extract_metadata(workbook):
    '''
    Hardcoded cell-location LUTs based on Pit_DDMMYY_SS_PP_V2.xlsx
    '''
    pit_sheet = workbook['PIT']
    # get shapely point as wkt
    try:
        lat = float(pit_sheet[PIT_LUT['Latitude']].value)
        lon = float(pit_sheet[PIT_LUT['Longitude']].value)
    except (TypeError, ValueError):
        pass
    try:
        geometry = dumps(Point([lon, lat]), rounding_precision=6)
    except (NameError, AttributeError):
        geometry = 'Unreadable'
    timestamp = parse_pit_datetime(
        date=pit_sheet[PIT_LUT['Date']].value,
        time=pit_sheet[PIT_LUT['Time']].value
    )
    # try to cast the slope and depth values to floats
    try:
        slope = float(pit_sheet[PIT_LUT['Slope']].value)
    except (TypeError, ValueError):
        slope = pit_sheet[PIT_LUT['Slope']].value
    try:
        total_depth = float(pit_sheet[PIT_LUT['Total_Depth']].value)
    except (TypeError, ValueError):
        total_depth = pit_sheet[PIT_LUT['Total_Depth']].value
    # setup the meta property for the given pit sheet
    meta_dict = {
        'location': pit_sheet[PIT_LUT['Location']].value,
        'timestamp': timestamp,
        'geometry': geometry,
        'surveyors': pit_sheet[PIT_LUT['Surveyors']].value,
        'site': pit_sheet[PIT_LUT['Site']].value,
        'pit_id': pit_sheet[PIT_LUT['Pit_ID']].value,
        'slope': slope,
        'total_depth': total_depth,
        'utm_zone': pit_sheet[PIT_LUT['UTM_Zone']].value,
        'comments': pit_sheet[PIT_LUT['Notes']].value
    }
    ## Also pull info from METADATA worksheet
    meta_sheet = workbook['METADATA']
    # only update the meta dict if there isn't already 
    # a value pulled from the PIT worksheet
    if meta_dict['location'] == None:
        meta_dict['location'] = meta_sheet[META_LUT['Location']].value
    if meta_dict['surveyors'] == None:
        meta_dict['surveyors'] = meta_sheet[META_LUT['Surveyors']].value
    if meta_dict['site'] == None:
        meta_dict['site'] =  meta_sheet[META_LUT['Site']].value
    if meta_dict['timestamp'] == None:
        meta_dict['timestamp'] = parse_pit_datetime(
            date=meta_sheet[META_LUT['Date']].value,
            time=meta_sheet[META_LUT['Time']].value
        )
    # append new information found on METADATA worksheet to meta_dict
    weather = meta_sheet[META_LUT['Weather']].value
    # precip
    precip = None
    for cell_idx, text_val in zip(META_LUT['Precip'], PRECIP):
        try:
            if int(meta_sheet[cell_idx].value) == 1:
                precip = text_val
                break
        except:
            pass
    # sky
    sky = None
    for cell_idx, text_val in zip(META_LUT['Sky'], SKY):
        try:
            if int(meta_sheet[cell_idx].value) == 1:
                sky = text_val
                break
        except:
            pass
    # wind_1
    wind_1 = None
    for cell_idx, text_val in zip(META_LUT['Wind_1'], WIND_1):
        try:
            if int(meta_sheet[cell_idx].value) == 1:
                wind_1 = text_val
                break
        except:
            pass
    # wind 2
    wind_2 = None
    for cell_idx, text_val in zip(META_LUT['Wind_2'], WIND_2):
        try:
            if int(meta_sheet[cell_idx].value) == 1:
                wind_2 = text_val
                break
        except:
            pass
    # ground condition
    ground_condition = None
    for cell_idx, text_val in zip(META_LUT['GroundCondition'], GRNDCON):
        try:
            if int(meta_sheet[cell_idx].value) == 1:
                ground_condition = text_val
                break
        except:
            pass
    # soil moisture
    soil_moisture = None
    for cell_idx, text_val in zip(META_LUT['SoilMoisture'], SOILMOIST):
        try:
            if int(meta_sheet[cell_idx].value) == 1:
                soil_moisture = text_val
                break
        except:
            pass
    # ground roughness
    ground_roughness = None
    for cell_idx, text_val in zip(META_LUT['GroundRoughness'], GRNDROUGH):
        try:
            if int(meta_sheet[cell_idx].value) == 1:
                ground_roughness = text_val
                break
        except:
            pass
    # ground vegetation **special case**
    veg_data = {}
    tmp_idx = 0
    for cell_idx, text_val in zip(META_LUT['GroundVegetation'], GRNDVEG):
        try:
            if int(meta_sheet[cell_idx].value) == 1:
                veg_class = text_val
                try:
                    veg_height = float(meta_sheet[META_LUT['HeightOfGroundVegetation'][tmp_idx]].value)
                except:
                    veg_height = 'Unreadable'
                veg_data[veg_class] = veg_height
            tmp_idx += 1
        except:
            pass
    # tree canopy
    tree_canopy = None
    for cell_idx, text_val in zip(META_LUT['TreeCanopy'], TREECANOPY):
        try:
            if int(meta_sheet[cell_idx].value) == 1:
                tree_canopy = text_val
                break
        except:
            pass
    meta_dict.update({
        'weather': weather,
        'precipitation': precip,
        'sky': sky,
        'wind_1': wind_1,
        'wind_2': wind_2,
        'ground_condition': ground_condition,
        'soil_moisture': soil_moisture,
        'ground_roughness': ground_roughness,
        'ground_vegetation': veg_data,
        'tree_canopy': tree_canopy
    })
    # Return the dict full of tasty infos
    return meta_dict


def extract_pit_data(sheet):
    '''
    Hardcoded cell locations for Pit_DDMMYY_SS_PP_V2.xlsx

    Pulls all the relevant rows into arrays/lists, then sets
    up a dataframe multi-indexed to the levels:
        1) Density
        2) Temperature
        3) Stratigraphy  ** 2 rows merged per cell value in xlsx sheet! **
        4) Specific Surface Area
    '''
    # openpyxl index slicing doesn't follow the rules of python list slicing...
    actual_max_row = sheet.max_row - 1
    # Density data
    dens_hagt = np.array([cell[0].value for cell in sheet[f'B10:B{actual_max_row}'] if cell[0].value not in ['Veg', 'NA']], dtype='float32')
    dens_hagb = np.array([cell[0].value for cell in sheet[f'D10:D{actual_max_row}'] if cell[0].value not in ['Veg', 'NA']], dtype='float32')
    dens_profA = np.array([cell[0].value for cell in sheet[f'E10:E{actual_max_row}'] if cell[0].value not in ['Veg', 'NA']], dtype='float32')
    dens_profB = np.array([cell[0].value for cell in sheet[f'F10:F{actual_max_row}'] if cell[0].value not in ['Veg', 'NA']], dtype='float32')
    density = pd.DataFrame(data={
        'HeightAboveGround_Top[cm]': pd.Series(dens_hagt, dtype='float32'),
        'HeightAboveGround_Bot[cm]': pd.Series(dens_hagb, dtype='float32'),
        'DensityProfile_A[kg/m^3]': pd.Series(dens_profA, dtype='float32'),
        'DensityProfile_B[kg/m^3]': pd.Series(dens_profB, dtype='float32')
    })
    # Temperature data
    temp_hag = np.array([cell[0].value for cell in sheet[f'G10:G{actual_max_row}'] if cell[0].value not in ['Veg', 'NA']], dtype='float32')
    temp_vals = np.array([cell[0].value for cell in sheet[f'H10:H{actual_max_row}'] if cell[0].value not in ['Veg', 'NA']], dtype='float32')
    temperature = pd.DataFrame(data={
        'HeightAboveGround[cm]': pd.Series(temp_hag, dtype='float32'),
        'Temperature[C]': pd.Series(temp_vals, dtype='float32')
    })
    # Stratigraphy data ** 2 rows merged per cell value in xlsx sheet! **
    strat_hagt = np.array([cell[0].value for cell in sheet[f'I10:I{actual_max_row}'][::2] if cell[0].value not in ['Veg', 'NA']], dtype='float32')
    strat_hagb = np.array([cell[0].value for cell in sheet[f'K10:K{actual_max_row}'][::2] if cell[0].value not in ['Veg', 'NA']], dtype='float32')
    strat_gs_min = np.array([cell[0].value for cell in sheet[f'L10:L{actual_max_row}'][::2] if cell[0].value not in ['Veg', 'NA']], dtype='float32')
    strat_gs_max = np.array([cell[0].value for cell in sheet[f'M10:M{actual_max_row}'][::2] if cell[0].value not in ['Veg', 'NA']], dtype='float32')
    strat_gs_avg = np.array([cell[0].value for cell in sheet[f'N10:N{actual_max_row}'][::2] if cell[0].value not in ['Veg', 'NA']], dtype='float32')
    strat_gt = [cell[0].value for cell in sheet[f'O10:O{actual_max_row}'][::2] if cell[0].value not in ['Veg', 'NA']]
    strat_gp = [cell[0].value for cell in sheet[f'P10:P{actual_max_row}'][::2] if cell[0].value not in ['Veg', 'NA']]
    strat_sw = [cell[0].value for cell in sheet[f'Q10:Q{actual_max_row}'][::2] if cell[0].value not in ['Veg', 'NA']]
    strat_notes = [cell[0].value for cell in sheet[f'V10:V{actual_max_row}'][::2] if cell[0].value not in ['Veg', 'NA']]
    stratigraphy = pd.DataFrame(data={
        'HeightAboveGround_Top[cm]': pd.Series(strat_hagt, dtype='float32'),
        'HeightAboveGround_Bot[cm]': pd.Series(strat_hagb, dtype='float32'),
        'GrainSize_min[mm]': pd.Series(strat_gs_min, dtype='float32'),
        'GrainSize_max[mm]': pd.Series(strat_gs_max, dtype='float32'),
        'GrainSize_mean[mm]': pd.Series(strat_gs_avg, dtype='float32'),
        'GrainType': pd.Series(strat_gt, dtype=object),
        'GrainPhoto': pd.Series(strat_gp, dtype=object),
        'SnowWetness': pd.Series(strat_sw, dtype=object),
        'Comments': pd.Series(strat_notes, dtype=object)
    })
    # need to re-index because of the 2-rows-per-cell-value in stratigraphy
    strat_idx = pd.Index([i for i in range(0, len(sheet[f'I10:I{actual_max_row}']), 2)])
    stratigraphy.index = strat_idx
    # Specific Surface Area data
    ssa_hagt = np.array([cell[0].value for cell in sheet[f'R10:R{actual_max_row}'] if cell[0].value not in ['Veg', 'NA']], dtype='float32')
    ssa_hagb = np.array([cell[0].value for cell in sheet[f'T10:T{actual_max_row}'] if cell[0].value not in ['Veg', 'NA']], dtype='float32')
    ssa_vals = np.array([cell[0].value for cell in sheet[f'U10:U{actual_max_row}'] if cell[0].value not in ['Veg', 'NA']], dtype='float32')
    specific_surface_area = pd.DataFrame(data={
        'HeightAboveGround_Top[cm]': pd.Series(ssa_hagt, dtype='float32'),
        'HeightAboveGround_Bot[cm]': pd.Series(ssa_hagb, dtype='float32'),
        'Val[V]': pd.Series(ssa_vals, dtype='float32')
    })
    # Set the <instance>.pit property to be a multi-indexed dataframe
    master_df = pd.concat({
        'density': density,
        'temperature': temperature,
        'stratigraphy': stratigraphy,
        'ssa': specific_surface_area,
    }, axis=1)
    # get rid of NaN rows 
    master_df = master_df.dropna(how='all').reset_index(drop=True)
    return master_df


def parse_pit_datetime(date, time):
    '''
    just in case the date fields are filled out by orangutans, 
    we take care when parsing the timestamp
    '''
    # sometimes the excel date field won't be preformatted as Date-type
    if isinstance(date, str):
        try:
            date = parse(date)
        except (ValueError, OverflowError):
            pass
    try:
        time = parse(time) # 12-hr with AM/PM works too!
    except TypeError:
        pass
    try:
        timestamp = pd.to_datetime(
            date.strftime('%Y-%m-%d') + 'T' + time.strftime('%H:%M:%S')
        ).isoformat()
    except (ValueError, TypeError):
        timestamp = 'Unreadable'
    return timestamp


if __name__ == '__main__':
    '''
    demo 
    '''
    test_xlsx = './demo/Pit_DDMMYY_SS_PP_V2.xlsx'
    foo = SnowPitSheet(test_xlsx)
    # fancy-pants string method overriding
    print(foo)
    print('\n')
    for key in foo.meta:
        print(f'{key}:    {foo.meta[key]}')
    print('\n')
    print(foo.data.info())
    print('\n')
    # possible to use .loc to get multiple columns belonging to different "levels"
    print(foo.data.loc[:, (slice(None), 'HeightAboveGround_Top[cm]')])
    print('\n')
    # can also treat each level ['ssa', 'density', 'temperature', 'stratigraphy'] as a property of the larger dataframe
    print(foo.data.density.columns)
    print('\n')
    print(foo.data.density['DensityProfile_A[kg/m^3]'].head(5))
